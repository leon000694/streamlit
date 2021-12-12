print()
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from google.oauth2.service_account import Credentials
import gspread
import sqlite3
from datetime import datetime

## 檢視各參數
#monthNo = int(datetime.now().strftime('%m'))-1 #當月No(mm)
monthNo = 10 #手動輸入月份
if monthNo == 0: monthNo = 12
print('M13 ', monthNo, '月份 重要業務數據~\n')

#抓預算月報重要數據 轉存M1路徑參數表
filepath_F = 'D:/python-training/1_project/M13月報/M13_110{}月報.xlsx'.format(monthNo) # 月報路徑
wb1 = load_workbook(filepath_F) 
ws1 = wb1.worksheets[0] 
char1 = get_column_letter(monthNo+2) #F檔當月欄名-----
data_F1 = ws1[char1 + str( 3)].value #F檔當月總預算
data_F2 = ws1[char1 + str(13)].value #F檔當月電話費
data_F3 = ws1[char1 + str(14)].value #F檔當月網路費
data_F4 = ws1[char1 + str(15)].value #F檔當月燃油費
data_F5 = ws1[char1 + str(16)].value #F檔用當月電費
data_F6 = ws1[char1 + str(17)].value #F檔用當月水費

filepath_O = 'D:/python-training/1_project/M13月報/M13重要業務數據表.xlsx' # 存M1
wb2 = load_workbook(filepath_O) 
ws2 = wb2.worksheets[0] 
char2 = get_column_letter(monthNo+2) #O檔當月欄名-----
ws2[char2 + str(3)].value = data_F1 #O檔總預算
ws2[char2 + str(4)].value = data_F2 #O檔電話費
ws2[char2 + str(5)].value = data_F3 #O檔網路費
ws2[char2 + str(6)].value = data_F4 #O檔燃油費
ws2[char2 + str(7)].value = data_F5 #O檔用電費
ws2[char2 + str(8)].value = data_F6 #O檔用水費

wb2.save(filepath_O) # 檔案回存 fileserver M1路徑

## 將各項資料append打包成df
combined = []
for i in [2,3,4,5,6,7,8]: 
	result = ws2[char2 + str(i)].value #取 M1檔案內單一數據
	print(ws2['A' + str(i)].value, 'is', result)
	combined.append(result) # combined為當月資料[list]
df_add = pd.DataFrame(combined) # 轉成DataFrame
df_add = np.transpose(df_add) # 縱資料轉置成橫向資料
df_add.columns = ['date','budget','telecom','network','fuel','electricity','waterpower']
print(df_add)

## (存至GoogleSheet) 將當月資料append累存
# 定義範圍及驗證gsp
scope = ['https://www.googleapis.com/auth/spreadsheets'] #google API應用範圍
creds = Credentials.from_service_account_file("D:/python-training/Tool/T_gsheet_Credentials.json", scopes=scope)
gs = gspread.authorize(creds)
gsheet = gs.open_by_url('https://docs.google.com/spreadsheets/d/1LJEUYcjOayfaKn1hiRbKRR3lPpj9unS1NvvPI2lcQZw/edit#gid=339883143') #---google-M13重要業務參數表
worksheet = gsheet.get_worksheet(0) 
worksheet.append_rows(df_add.values.tolist()) #將當月資料附加上去

## (存至桌機 SQLite3 DB) 將當月資料append累存"""
conn = sqlite3.connect('D:/python-training/1_project/M13月報/M13-parameter.db') #連接db資料庫/無則建立------
cursor = conn.cursor()
cursor.execute("CREATE TABLE IF NOT EXISTS Parameter('date','budget','telecom','network','fuel','electricity','waterpower')") # 
conn.commit()
df_add.to_sql('Parameter', conn, if_exists='append', index=False) #將df_add資料累存db檔
conn.close()


print('\n程式執行完成...\n')
# https://www.learncodewithmike.com/2021/06/pandas-and-google-sheets.html
# Google API Developer Console首頁 https://console.cloud.google.com/apis
# https://www.learncodewithmike.com/2020/11/python-pandas-dataframe-tutorial.html