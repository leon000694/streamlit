print()
from openpyxl import Workbook
from openpyxl import load_workbook
import random

fileName = "temp.xlsx" # ================================================ 設定檔名
wb = Workbook()
ws = wb.active
print("創建的活頁簿名: ", ws.title)

ws.title = "Space-X" # 修改sheet的名稱
ws.sheet_properties.tabColor = "FF0000" # 修改下面tab的顏色

# 創建新Sheet,可以控制位置,都表示是在這個位置之前插入
ws1 = wb.create_sheet("MySheet1",1)
ws2 = wb.create_sheet("MySheet2",2)
ws3 = wb.create_sheet("MySheet3",3)

# 遍歷所有的sheet
for sheetname in wb.sheetnames:
    print("所有Sheet名/順序:", sheetname)

# 選中的sheet
ws3 = wb["MySheet3"]
print("選中的sheet-choiced:", ws3.title)
# 修改active sheet名
wb.active = wb["MySheet3"]
ws3 = wb.active
print("當前的sheet-actived:", ws3.title)

# Section2 Cell層面 ====================================================== 賦值
ws3["A1"] = 1
a1 = ws3["A1"]
print("A1的值:", a1.value,'\n')

cell_range = ws3["A1":"C5"]
for row in cell_range:
    for cell in row:
        cell.value = random.randint(1,100) # ============================ 產值        
        print("4.  列{0} 欄{1} 值 {2}".format(cell.row,cell.column,cell.value))
        # (A1 B1 C1/A2 B2 C2/A3 B3 C3/A4 B4 C4/A5 B5 C5)
print()
 
# # iter_rows 遍歷行 ====================================================== 取值
# for row in ws3.iter_rows(min_row=1, max_row=2, min_col=2, max_col=3, values_only=False):
#      for cell in row:
#         print("5.iter_row 行{0}列{1}值{2}".format(cell.row,cell.column,cell.value))   

# # iter_cols 遍歷欄
# for col in ws3.iter_cols(min_col=2, max_col=3, min_row=1,max_row=2, values_only=False):
#     for cell in col:
#         print("iter_col 行{0}列{1}值{2}".format(cell.row,cell.column,cell.value))
# print()

# # 只想值遍歷 ============================================================ 取值
# for row in ws3.values:
#     print("值遍歷:", row)
# print()

# # iter_row和 iter_cols也支持值遍歷
# for row in ws3.iter_rows(min_row=1,min_col=2,max_row=2,max_col=3,values_only=True):
#     print("iter_row 值遍歷: ", row)
# print()    

# #保存
# wb.save(fileName) # ==================================================== 存檔

# # 加載已存在
# wb = load_workbook(fileName)
# ws = wb.active
# for row in ws.values:
#     print("已存在:", row)
# print()

# # 亂數存值，找到最大值, 最小值 =========================================== 取最大 最小值 和
# wb = load_workbook(fileName)
# ws = wb.create_sheet("Exercise")
# cell_range = ws["A1":"D6"]
# for row in cell_range:
#     for cell in row:
#         cell.value = random.randint(1,100)
# maxValue = 0
# minValue = 101
# sum = 0
# for row in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=4,values_only=True):
#     for val in row:
#         if val > maxValue:
#             maxValue = val
#         if val < minValue:
#             minValue = val
#         sum += val
# ws["C7"] = "maxValue"
# ws["D7"] = maxValue
# ws["C8"] = "minValue"
# ws["D8"] = minValue
# ws["C9"] = "sum"
# ws["D9"] = sum

wb.save(fileName)

print()
# 教學網址: https://www.youtube.com/watch?v=4fjWuS47UQA