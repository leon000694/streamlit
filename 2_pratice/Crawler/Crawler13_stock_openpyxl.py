import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import requests

def export(self, stocks):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("Yahoo股市", 0)
    response = requests.get("https://tw.stock.yahoo.com/q/q?s=2451")
    soup = BeautifulSoup(response.text,"lxml")
    tables = soup.find_all("table")[2]
    ths = tables.find_all("th")[0:11]
    titles = ("資料日期",) + tuple(th.getText() for th in ths)
    sheet.append(titles)

    for index, stock in enumerate(stocks):
        sheet.append(stock)
        print("紅") #儲存格字體紅色
        if "" in stock[6]:
            sheet.cell(row=index+2, column=7).font = Font(color='FFoooo') 
        elif "" in stock[6]:
            sheet.cell(row=index+2, column=7).font = Font(color='00A600')
            print("綠") #儲存格字體顯示綠色
            
    wb.save("yahoostock.xlsx")

stock = Stock('2451', '2454', '2369')
stock.export(stock.scrape())
